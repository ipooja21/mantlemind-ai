import os
import time
import io
from datetime import datetime
from typing import Optional
import pandas as pd
import numpy as np
from loguru import logger

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from core.config import REPORT_OUTPUT_DIR


try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("[ReportAgent] reportlab not installed — PDF disabled")


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("[ReportAgent] openpyxl not installed — Excel disabled")



if PDF_AVAILABLE:
    MANTLE_TEAL   = colors.HexColor("#00C2B2")
    MANTLE_DARK   = colors.HexColor("#0A1628")
    MANTLE_GREY   = colors.HexColor("#8A9BB0")
    ACCENT_GREEN  = colors.HexColor("#00E676")
    ACCENT_ORANGE = colors.HexColor("#FF9800")
    ACCENT_RED    = colors.HexColor("#F44336")
    WHITE         = colors.white


class ReportAgent:
    
    def __init__(self):
        os.makedirs(REPORT_OUTPUT_DIR, exist_ok=True)
        logger.success("[ReportAgent] Initialised")

    
    def generate_pdf(
        self,
        df: pd.DataFrame,
        anomaly_summary: dict,
        whale_alerts: list,
        alpha: dict,
        network_stats: dict,
    ) -> Optional[str]:
        
        if not PDF_AVAILABLE:
            logger.error("[ReportAgent] reportlab not available")
            return None

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(REPORT_OUTPUT_DIR, f"MantleMind_Report_{ts}.pdf")

        doc = SimpleDocTemplate(
            path, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm,
            topMargin=2*cm, bottomMargin=2*cm,
        )

        styles = getSampleStyleSheet()
        story = []

        
        title_style = ParagraphStyle(
            "title", parent=styles["Title"],
            fontSize=24, textColor=MANTLE_TEAL, spaceAfter=6,
        )
        sub_style = ParagraphStyle(
            "sub", parent=styles["Normal"],
            fontSize=11, textColor=MANTLE_GREY, spaceAfter=20,
        )
        story.append(Paragraph("MantleMind AI", title_style))
        story.append(Paragraph(
            f"Intelligence Report — {datetime.now().strftime('%d %B %Y, %H:%M UTC')}",
            sub_style,
        ))
        story.append(HRFlowable(width="100%", thickness=2, color=MANTLE_TEAL))
        story.append(Spacer(1, 0.4*cm))

        
        h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=MANTLE_TEAL, spaceBefore=12)
        body = ParagraphStyle("body", parent=styles["Normal"], fontSize=10, leading=16)

        story.append(Paragraph("Executive Summary", h2))
        alpha_score = alpha.get("alpha_score", 50)
        signal = alpha.get("signal", "NEUTRAL")
        summary_text = (
            f"Mantle Network analysis covering <b>{len(df)}</b> transactions "
            f"across <b>{df['block_number'].nunique() if 'block_number' in df.columns else 'N/A'}</b> blocks. "
            f"Current Alpha Score: <b>{alpha_score}/100</b> ({signal}). "
            f"Anomaly rate: <b>{anomaly_summary.get('anomaly_rate', 0):.1f}%</b>. "
            f"Whale wallets detected: <b>{len(whale_alerts)}</b>."
        )
        story.append(Paragraph(summary_text, body))
        story.append(Spacer(1, 0.3*cm))

        
        story.append(Paragraph("Network Statistics", h2))
        net_data = [
            ["Metric", "Value"],
            ["Status", "LIVE" if network_stats.get("connected") else "DEMO"],
            ["Latest Block", f"{network_stats.get('latest_block', 'N/A'):,}"],
            ["Gas Price", f"{network_stats.get('gas_price_gwei', 0):.6f} Gwei"],
            ["Chain ID", str(network_stats.get("chain_id", 5000))],
            ["Total Transactions (batch)", str(len(df))],
        ]
        story.append(self._make_table(net_data))
        story.append(Spacer(1, 0.4*cm))

        
        story.append(Paragraph("Anomaly Detection Summary", h2))
        anom_data = [
            ["Metric", "Value"],
            ["Total Anomalies", str(anomaly_summary.get("total_anomalies", 0))],
            ["Anomaly Rate", f"{anomaly_summary.get('anomaly_rate', 0):.2f}%"],
            ["Critical Transactions", str(anomaly_summary.get("critical_count", 0))],
            ["Max Anomaly Score", str(anomaly_summary.get("max_anomaly_score", 0))],
            ["Avg Anomaly Score", str(anomaly_summary.get("avg_anomaly_score", 0))],
        ]
        story.append(self._make_table(anom_data))
        story.append(Spacer(1, 0.4*cm))

        
        story.append(Paragraph("Whale Activity", h2))
        if whale_alerts:
            whale_data = [["Address", "Tier", "Pattern", "Volume (MNT)", "Score"]]
            for w in whale_alerts[:10]:
                whale_data.append([
                    str(w.get("address", ""))[:18] + "…",
                    w.get("tier", ""),
                    w.get("pattern", ""),
                    f"{w.get('total_mnt', 0):,.0f}",
                    str(w.get("score", 0)),
                ])
            story.append(self._make_table(whale_data, header_row=True))
        else:
            story.append(Paragraph("No whale activity detected in this period.", body))
        story.append(Spacer(1, 0.4*cm))

        
        story.append(Paragraph("Alpha Signal", h2))
        alpha_data = [
            ["Component", "Score"],
            ["Alpha Score", f"{alpha.get('alpha_score', 0)}/100"],
            ["Signal", alpha.get("signal", "")],
            ["Whale Signal", str(alpha.get("components", {}).get("whale_signal", 0))],
            ["Anomaly Penalty", str(alpha.get("components", {}).get("anomaly_penalty", 0))],
            ["Volume Trend", str(alpha.get("components", {}).get("volume_trend", 0))],
            ["Gas Momentum", str(alpha.get("components", {}).get("gas_momentum", 0))],
            ["Contract Activity", str(alpha.get("components", {}).get("contract_activity", 0))],
        ]
        story.append(self._make_table(alpha_data))

        narrative = alpha.get("narrative", "")
        if narrative:
            story.append(Spacer(1, 0.3*cm))
            story.append(Paragraph("AI Narrative Insight", h2))
            
            clean = narrative.replace("**", "").replace("*", "").replace("🚀","").replace("📈","").replace("🐋","").replace("⚠️","").replace("✅","").replace("💡","").replace("📊","")
            for line in clean.split("\n"):
                if line.strip():
                    story.append(Paragraph(line.strip(), body))

        
        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=MANTLE_GREY))
        footer_style = ParagraphStyle("footer", parent=styles["Normal"], fontSize=8,
                                       textColor=MANTLE_GREY, alignment=TA_CENTER)
        story.append(Paragraph(
            f"Generated by MantleMind AI · {datetime.now().strftime('%Y-%m-%d %H:%M UTC')} · "
            f"Mantle Network (Chain ID 5000)",
            footer_style,
        ))

        doc.build(story)
        logger.success(f"[ReportAgent] PDF generated: {path}")
        return path

    
    def generate_excel(
        self,
        df: pd.DataFrame,
        anomaly_summary: dict,
        whale_alerts: list,
        alpha: dict,
        network_stats: dict,
    ) -> Optional[str]:
        
        if not EXCEL_AVAILABLE:
            logger.error("[ReportAgent] openpyxl not available")
            return None

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(REPORT_OUTPUT_DIR, f"MantleMind_Report_{ts}.xlsx")
        wb = Workbook()

        
        ws = wb.active
        ws.title = "Summary"
        self._write_summary_sheet(ws, anomaly_summary, whale_alerts, alpha, network_stats, len(df))

        
        if not df.empty:
            ws2 = wb.create_sheet("Transactions")
            self._write_df_sheet(ws2, df.head(1000))

        
        if "is_anomaly" in df.columns:
            anom_df = df[df["is_anomaly"] == True].copy() if not df.empty else pd.DataFrame()
            if not anom_df.empty:
                ws3 = wb.create_sheet("Anomalies")
                self._write_df_sheet(ws3, anom_df)

        
        if whale_alerts:
            ws4 = wb.create_sheet("Whale Activity")
            whale_df = pd.DataFrame(whale_alerts)[["address","tier","pattern","total_mnt","score","tx_count"]]
            self._write_df_sheet(ws4, whale_df)

        
        ws5 = wb.create_sheet("Alpha Signal")
        self._write_alpha_sheet(ws5, alpha)

        wb.save(path)
        logger.success(f"[ReportAgent] Excel generated: {path}")
        return path

    
    def _write_summary_sheet(self, ws, anomaly_summary, whale_alerts, alpha, network_stats, tx_count):
        teal_fill = PatternFill("solid", fgColor="00C2B2")
        dark_fill = PatternFill("solid", fgColor="0A1628")
        bold_white = Font(bold=True, color="FFFFFF", size=12)
        bold_teal = Font(bold=True, color="00C2B2", size=11)

        ws["A1"] = "MantleMind AI — Intelligence Report"
        ws["A1"].font = Font(bold=True, size=16, color="00C2B2")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}"
        ws["A2"].font = Font(italic=True, color="8A9BB0")
        ws.merge_cells("A1:F1")

        rows = [
            ("NETWORK", ""),
            ("Status", "LIVE" if network_stats.get("connected") else "DEMO"),
            ("Latest Block", network_stats.get("latest_block", "N/A")),
            ("Gas Price (Gwei)", network_stats.get("gas_price_gwei", 0)),
            ("Chain ID", network_stats.get("chain_id", 5000)),
            ("", ""),
            ("ANOMALIES", ""),
            ("Total Transactions", tx_count),
            ("Total Anomalies", anomaly_summary.get("total_anomalies", 0)),
            ("Anomaly Rate (%)", anomaly_summary.get("anomaly_rate", 0)),
            ("Critical Transactions", anomaly_summary.get("critical_count", 0)),
            ("", ""),
            ("WHALES", ""),
            ("Total Whale Wallets", len(whale_alerts)),
            ("Mega Whales", sum(1 for w in whale_alerts if w.get("tier") == "MEGA")),
            ("Total Volume (MNT)", sum(w.get("total_mnt", 0) for w in whale_alerts)),
            ("", ""),
            ("ALPHA SIGNAL", ""),
            ("Alpha Score", alpha.get("alpha_score", 0)),
            ("Signal", alpha.get("signal", "NEUTRAL")),
        ]

        for i, (key, val) in enumerate(rows, start=4):
            ws.cell(row=i, column=1, value=key)
            ws.cell(row=i, column=2, value=val)
            if val == "" and key in ("NETWORK", "ANOMALIES", "WHALES", "ALPHA SIGNAL"):
                ws.cell(row=i, column=1).font = bold_teal

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25

    def _write_df_sheet(self, ws, df: pd.DataFrame):
        teal_fill = PatternFill("solid", fgColor="00C2B2")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = teal_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row, start=1):
                v = val
                if isinstance(v, (np.integer,)): v = int(v)
                elif isinstance(v, (np.floating,)): v = round(float(v), 6)
                elif pd.isna(v) if not isinstance(v, str) else False: v = ""
                ws.cell(row=row_idx, column=col_idx, value=v)

        for i, col in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(col)) + 4)

    def _write_alpha_sheet(self, ws, alpha: dict):
        ws["A1"] = "Alpha Signal Components"
        ws["A1"].font = Font(bold=True, size=14, color="00C2B2")
        components = alpha.get("components", {})
        rows = [
            ("Alpha Score", alpha.get("alpha_score", 0)),
            ("Signal", alpha.get("signal", "")),
            ("", ""),
            ("Whale Signal (×0.35)", components.get("whale_signal", 0)),
            ("Anomaly Penalty (×0.25)", components.get("anomaly_penalty", 0)),
            ("Volume Trend (×0.20)", components.get("volume_trend", 0)),
            ("Gas Momentum (×0.10)", components.get("gas_momentum", 0)),
            ("Contract Activity (×0.10)", components.get("contract_activity", 0)),
        ]
        for i, (k, v) in enumerate(rows, start=3):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)

    
    def _make_table(self, data: list, header_row: bool = True) -> "Table":
        t = Table(data, hAlign="LEFT")
        style = [
            ("FONTSIZE",    (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F5F7FA"), colors.white]),
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#DEE3EA")),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ]
        if header_row:
            style += [
                ("BACKGROUND",  (0, 0), (-1, 0), MANTLE_TEAL),
                ("TEXTCOLOR",   (0, 0), (-1, 0), WHITE),
                ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        t.setStyle(TableStyle(style))
        return t

    
    def generate_pdf_bytes(self, df, anomaly_summary, whale_alerts, alpha, network_stats) -> bytes:
        
        path = self.generate_pdf(df, anomaly_summary, whale_alerts, alpha, network_stats)
        if path and os.path.exists(path):
            with open(path, "rb") as f:
                return f.read()
        return b""

    def generate_excel_bytes(self, df, anomaly_summary, whale_alerts, alpha, network_stats) -> bytes:
        
        path = self.generate_excel(df, anomaly_summary, whale_alerts, alpha, network_stats)
        if path and os.path.exists(path):
            with open(path, "rb") as f:
                return f.read()
        return b""



if __name__ == "__main__":
    from agents.agent1_data_collector import DataCollectorAgent
    from agents.agent2_anomaly_detector import AnomalyDetectorAgent
    from agents.agent3_whale_tracker import WhaleTrackerAgent
    from agents.agent4_alpha_generator import AlphaGeneratorAgent

    c = DataCollectorAgent(); df = c.fetch_latest_blocks(10)
    d = AnomalyDetectorAgent(); df = d.score(df); asum = d.get_summary(df)
    wt = WhaleTrackerAgent(); df, walerts = wt.detect_whales(df)
    ag = AlphaGeneratorAgent(); alpha = ag.generate(df, asum, walerts)
    ns = c.get_network_stats()

    r = ReportAgent()
    print(r.generate_pdf(df, asum, walerts, alpha, ns))
    print(r.generate_excel(df, asum, walerts, alpha, ns))
